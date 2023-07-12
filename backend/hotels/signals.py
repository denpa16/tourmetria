from django.dispatch import receiver
from django.db.models import signals
from openpyxl import load_workbook

from hotels.models import Hotel
from rooms.models import Room, RoomImage, RoomFeature


@receiver(signal=signals.pre_save, sender=Hotel)
def get_hotel_rooms(instance: Hotel, **kwargs):
    """
    Получение номеров отеля

    """
    if Room.objects.filter(hotel=instance).count() == 0:

        def parse(sheet_obj, row_number):
            """
            Парсер данных в строке файла

            """
            property_fields = {
                "hotel": sheet_obj.cell(row=row_number, column=1).value,
                "name": sheet_obj.cell(row=row_number, column=2).value,
                "capacity": sheet_obj.cell(row=row_number, column=3).value,
                "description": sheet_obj.cell(row=row_number, column=4).value,
                "images": sheet_obj.cell(row=row_number, column=5).value,
                "features": sheet_obj.cell(row=row_number, column=6).value,
            }

            return property_fields

        file_path = "hotels/newfile.xlsx"
        wb_obj = load_workbook(file_path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        for i in range(2, max_row + 1):
            room_data = parse(sheet_obj, i)
            room, _ = Room.objects.get_or_create(
                hotel=instance,
                name=room_data["name"],
                description=room_data["description"],
                capacity=room_data["capacity"],
            )
            images = room_data["images"].split(", ")
            features = room_data["features"].split(", ")
            for i in images:
                RoomImage.objects.create(room=room, url=i)
            room_features = list()
            for f in features:
                feature, _ = RoomFeature.objects.get_or_create(name=f)
                room_features.append(feature)
            room.features.set(room_features)
